import json
import boto3
import os
import re
import logging
from urllib.parse import unquote_plus
import tempfile
from io import BytesIO
import time
from botocore.exceptions import ClientError

# Configure logging first
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Document processing libraries with better error handling
try:
    from docx import Document
    from docx.shared import Inches
    DOCX_AVAILABLE = True
except ImportError as e:
    logger.warning(f"DOCX import failed: {str(e)}")
    Document = None
    DOCX_AVAILABLE = False

try:
    import pypdf
    PYPDF_AVAILABLE = True
except ImportError as e:
    logger.warning(f"pypdf import failed: {str(e)}")
    pypdf = None
    PYPDF_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    logger.warning(f"openpyxl import failed: {str(e)}")
    load_workbook = None
    OPENPYXL_AVAILABLE = False

# Initialize AWS clients
s3 = boto3.client('s3')

# Configuration constants
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB limit
MAX_CONFIG_SIZE = 1 * 1024 * 1024  # 1MB limit for config
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'docx', 'doc', 'xlsx', 'xls'}
MAX_REPLACEMENTS = 100  # Maximum number of replacement rules
BATCH_SIZE = 5  # Maximum files to process in one batch
BATCH_TIMEOUT = 45  # Maximum seconds for batch processing

# Retry configuration
MAX_RETRIES = 3
BASE_BACKOFF = 1  # seconds
MAX_BACKOFF = 30  # seconds

# Environment variables
INPUT_BUCKET = os.environ['INPUT_BUCKET']
OUTPUT_BUCKET = os.environ['OUTPUT_BUCKET']
QUARANTINE_BUCKET = os.environ['QUARANTINE_BUCKET']
CONFIG_BUCKET = os.environ['CONFIG_BUCKET']

# Global cache for config
_config_cache = None
_config_last_modified = None

# Default fallback patterns if config not available
DEFAULT_REPLACEMENTS = [
    {"find": "REPLACE_CLIENT_NAME", "replace": "Company X"},
    {"find": "ACME Corporation", "replace": "[REDACTED]"},
    {"find": "TechnoSoft", "replace": "[REDACTED]"}
]

def validate_file(bucket, key):
    """
    Validate file before processing
    """
    try:
        # Check file extension
        file_ext = key.lower().split('.')[-1]
        if file_ext not in ALLOWED_EXTENSIONS:
            raise ValueError(f"Unsupported file type: {file_ext}. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}")
        
        # Check file size
        response = s3.head_object(Bucket=bucket, Key=key)
        file_size = response['ContentLength']
        
        if file_size > MAX_FILE_SIZE:
            raise ValueError(f"File too large: {file_size} bytes. Maximum allowed: {MAX_FILE_SIZE} bytes")
        
        if file_size == 0:
            raise ValueError("File is empty")
        
        return True
    except s3.exceptions.NoSuchKey:
        raise ValueError(f"File not found: {key}")
    except Exception as e:
        logger.error(f"File validation error: {str(e)}")
        raise

def validate_config(config):
    """
    Validate configuration structure and content
    """
    if not isinstance(config, dict):
        raise ValueError("Configuration must be a JSON object")
    
    # Check required fields
    if 'replacements' not in config:
        raise ValueError("Configuration must contain 'replacements' field")
    
    replacements = config.get('replacements', [])
    
    if not isinstance(replacements, list):
        raise ValueError("'replacements' must be an array")
    
    if len(replacements) > MAX_REPLACEMENTS:
        raise ValueError(f"Too many replacement rules: {len(replacements)}. Maximum allowed: {MAX_REPLACEMENTS}")
    
    # Validate each replacement rule
    for idx, rule in enumerate(replacements):
        if not isinstance(rule, dict):
            raise ValueError(f"Replacement rule {idx} must be an object")
        
        if 'find' not in rule:
            raise ValueError(f"Replacement rule {idx} missing required 'find' field")
        
        if not rule.get('find'):
            raise ValueError(f"Replacement rule {idx} has empty 'find' field")
        
        if 'replace' not in rule:
            logger.warning(f"Replacement rule {idx} missing 'replace' field, using '[REDACTED]'")
    
    return True

def lambda_handler(event, context):
    """
    Main Lambda handler for document processing with batch support
    """
    results = []
    start_time = time.time()
    
    try:
        # Load configuration once for the entire batch
        config = get_redaction_config()
        validate_config(config)
        
        # Get files to process
        files_to_process = []
        for record in event['Records']:
            bucket = record['s3']['bucket']['name']
            key = unquote_plus(record['s3']['object']['key'])
            files_to_process.append((bucket, key))
        
        # Log batch start
        logger.info(json.dumps({
            'event': 'BATCH_START',
            'batch_size': len(files_to_process),
            'request_id': context.aws_request_id
        }))
        
        # Process files in batches
        processed_count = 0
        for i in range(0, len(files_to_process), BATCH_SIZE):
            batch = files_to_process[i:i + BATCH_SIZE]
            
            # Check timeout
            elapsed = time.time() - start_time
            if elapsed > BATCH_TIMEOUT:
                logger.warning(json.dumps({
                    'event': 'BATCH_TIMEOUT',
                    'processed': processed_count,
                    'remaining': len(files_to_process) - processed_count,
                    'elapsed': elapsed
                }))
                break
            
            # Process batch
            batch_results = process_file_batch(batch, config, context)
            results.extend(batch_results)
            processed_count += len(batch_results)
        
        # Log batch completion
        logger.info(json.dumps({
            'event': 'BATCH_COMPLETE',
            'processed': processed_count,
            'total': len(files_to_process),
            'elapsed': time.time() - start_time,
            'request_id': context.aws_request_id
        }))
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'message': 'Batch processing completed',
                'processed': processed_count,
                'total': len(files_to_process),
                'results': results
            })
        }
        
    except Exception as e:
        logger.error(json.dumps({
            'event': 'HANDLER_ERROR',
            'error': str(e),
            'request_id': context.aws_request_id
        }))
        
        return {
            'statusCode': 500,
            'body': json.dumps({
                'error': str(e),
                'results': results
            })
        }

def process_file_batch(files_batch, config, context):
    """
    Process a batch of files
    """
    results = []
    
    for bucket, key in files_batch:
        result = {
            'file': key,
            'status': 'PROCESSING',
            'timestamp': context.aws_request_id
        }
        
        try:
            # Validate file before processing
            validate_file(bucket, key)
            
            logger.info(json.dumps({
                'event': 'PROCESSING_START',
                'file': key,
                'bucket': bucket,
                'request_id': context.aws_request_id
            }))
            
            # Process document based on file type
            file_ext = key.lower().split('.')[-1]
            
            if file_ext == 'txt':
                success = process_text_file(bucket, key, config)
            elif file_ext == 'pdf':
                success = process_pdf_file(bucket, key, config)
            elif file_ext in ['docx', 'doc']:
                success = process_docx_file(bucket, key, config)
            elif file_ext in ['xlsx', 'xls']:
                success = process_xlsx_file(bucket, key, config)
            else:
                quarantine_document(bucket, key, f"Unsupported file type: {file_ext}")
                success = False
            
            result['status'] = 'SUCCESS' if success else 'QUARANTINED'
            
            logger.info(json.dumps({
                'event': 'PROCESSING_COMPLETE',
                'file': key,
                'status': result['status'],
                'request_id': context.aws_request_id
            }))
            
        except Exception as e:
            logger.error(json.dumps({
                'event': 'PROCESSING_ERROR',
                'file': key,
                'error': str(e),
                'request_id': context.aws_request_id
            }))
            
            # Move to quarantine on error
            try:
                quarantine_document(bucket, key, f"Processing error: {str(e)}")
            except Exception as qe:
                logger.error(f"Failed to quarantine {key}: {str(qe)}")
            
            result['status'] = 'ERROR'
            result['error'] = str(e)
        
        results.append(result)
    
    return results

def exponential_backoff_retry(func, *args, **kwargs):
    """
    Retry function with exponential backoff
    """
    last_exception = None
    
    for attempt in range(MAX_RETRIES):
        try:
            return func(*args, **kwargs)
        except ClientError as e:
            last_exception = e
            error_code = e.response['Error']['Code']
            
            # Don't retry on permanent errors
            if error_code in ['NoSuchKey', 'NoSuchBucket', 'AccessDenied', 'InvalidRequest']:
                raise
            
            if attempt < MAX_RETRIES - 1:
                backoff = min(BASE_BACKOFF * (2 ** attempt), MAX_BACKOFF)
                logger.warning(f"Retry {attempt + 1}/{MAX_RETRIES} after {backoff}s for {func.__name__}")
                time.sleep(backoff)
            else:
                raise
        except Exception as e:
            # For non-ClientError exceptions, fail immediately
            raise
    
    raise last_exception

def download_document(bucket, key):
    """Download document from S3 with retry logic"""
    def _download():
        response = s3.get_object(Bucket=bucket, Key=key)
        return response['Body'].read()
    
    try:
        return exponential_backoff_retry(_download)
    except Exception as e:
        logger.error(f"Error downloading {key} after {MAX_RETRIES} attempts: {str(e)}")
        raise

def get_redaction_config():
    """Load redaction configuration from S3 with caching"""
    global _config_cache, _config_last_modified
    
    try:
        # Check if config file exists and get last modified time
        response = s3.head_object(Bucket=CONFIG_BUCKET, Key='config.json')
        last_modified = response['LastModified']
        config_size = response['ContentLength']
        
        # Validate config file size
        if config_size > MAX_CONFIG_SIZE:
            raise ValueError(f"Configuration file too large: {config_size} bytes")
        
        # Return cached config if it's still current
        if _config_cache and _config_last_modified and last_modified == _config_last_modified:
            return _config_cache
        
        # Download and parse config
        config_response = s3.get_object(Bucket=CONFIG_BUCKET, Key='config.json')
        config_content = config_response['Body'].read()
        
        try:
            config_data = json.loads(config_content)
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in configuration: {str(e)}")
            raise ValueError("Configuration file contains invalid JSON")
        
        # Cache the config
        _config_cache = config_data
        _config_last_modified = last_modified
        
        logger.info(f"Loaded redaction config with {len(config_data.get('replacements', []))} rules")
        return config_data
        
    except s3.exceptions.NoSuchKey:
        logger.warning("No config.json found, using default replacements")
        return {'replacements': DEFAULT_REPLACEMENTS, 'case_sensitive': False}
    except Exception as e:
        logger.error(f"Error loading config: {str(e)}, using defaults")
        return {'replacements': DEFAULT_REPLACEMENTS, 'case_sensitive': False}

def apply_redaction_rules(content, config):
    """Apply redaction rules to text content"""
    try:
        processed_content = content
        redactions_made = []
        
        replacements = config.get('replacements', DEFAULT_REPLACEMENTS)
        case_sensitive = config.get('case_sensitive', False)
        flags = 0 if case_sensitive else re.IGNORECASE
        
        for rule in replacements:
            find_text = rule.get('find', '')
            replace_text = rule.get('replace', '[REDACTED]')
            
            if not find_text:
                continue
                
            # Count matches before replacement
            matches = re.findall(re.escape(find_text), processed_content, flags=flags)
            if matches:
                redactions_made.extend([(find_text, len(matches))])
                processed_content = re.sub(re.escape(find_text), replace_text, processed_content, flags=flags)
        
        if redactions_made:
            logger.info(f"Applied redactions: {redactions_made}")
        
        return processed_content, len(redactions_made) > 0
        
    except Exception as e:
        logger.error(f"Error applying redaction rules: {str(e)}")
        return content, False

def apply_filename_redaction(filename, config):
    """Apply redaction rules to file names"""
    try:
        # Extract base name and extension
        base_name = filename.rsplit('.', 1)[0] if '.' in filename else filename
        extension = filename.rsplit('.', 1)[1] if '.' in filename else ''
        
        # Apply redaction to base name only (preserve extension)
        processed_name, redacted = apply_redaction_rules(base_name, config)
        
        # Reconstruct filename with extension
        if extension:
            processed_filename = f"{processed_name}.{extension}"
        else:
            processed_filename = processed_name
            
        if redacted:
            logger.info(f"Applied filename redaction: {filename} -> {processed_filename}")
            
        return processed_filename, redacted
        
    except Exception as e:
        logger.error(f"Error applying filename redaction: {str(e)}")
        return filename, False

def upload_processed_document(key, content, metadata=None, config=None):
    """Upload processed document to output bucket with retry logic"""
    # Apply filename redaction if config provided
    if config:
        redacted_key, filename_redacted = apply_filename_redaction(key, config)
        processed_key = f"processed/{redacted_key}"
        if metadata is None:
            metadata = {}
        metadata['filename_redacted'] = str(filename_redacted)
    else:
        processed_key = f"processed/{key}"
    
    base_metadata = {
        'processing-status': 'completed',
        'original-key': key
    }
    
    if metadata:
        base_metadata.update(metadata)
    
    # Handle both bytes and file-like objects
    if hasattr(content, 'read'):
        body = content.read()
    else:
        body = content
    
    def _upload():
        s3.put_object(
            Bucket=OUTPUT_BUCKET,
            Key=processed_key,
            Body=body,
            ServerSideEncryption='AES256',
            Metadata=base_metadata
        )
    
    try:
        exponential_backoff_retry(_upload)
        logger.info(f"Uploaded processed document: s3://{OUTPUT_BUCKET}/{processed_key}")
        return True
    except Exception as e:
        logger.error(f"Error uploading processed document after {MAX_RETRIES} attempts: {str(e)}")
        raise

def process_text_file(bucket, key, config):
    """Process text files"""
    try:
        # Download document
        document_content = download_document(bucket, key)
        text_content = document_content.decode('utf-8')
        
        # Apply redaction rules
        processed_content, redacted = apply_redaction_rules(text_content, config)
        
        # Upload processed document
        upload_processed_document(key, processed_content.encode('utf-8'), 
                                {'redacted': str(redacted)}, config)
        
        return True
        
    except Exception as e:
        logger.error(f"Error processing text file {key}: {str(e)}")
        raise

def process_pdf_file(bucket, key, config):
    """Process PDF files - extract text and output as redacted text file"""
    if not pypdf:
        raise ImportError("pypdf library not available")
    
    try:
        # Download document
        document_content = download_document(bucket, key)
        
        with tempfile.NamedTemporaryFile(suffix='.pdf') as temp_file:
            temp_file.write(document_content)
            temp_file.flush()
            
            # Read PDF and extract all text
            reader = pypdf.PdfReader(temp_file.name)
            
            all_text = []
            for page in reader.pages:
                # Extract text from page
                text = page.extract_text()
                if text.strip():
                    all_text.append(text)
            
            # Combine all text
            full_text = '\n\n'.join(all_text)
            
            # Apply redaction rules
            processed_text, redacted = apply_redaction_rules(full_text, config)
            
            # Save as text file (change extension to .txt)
            text_key = key.rsplit('.', 1)[0] + '.txt'
            
            # Upload processed document as text
            upload_processed_document(text_key, processed_text.encode('utf-8'), 
                                    {'redacted': str(redacted), 'converted_from': 'pdf'}, config)
            
            return True
            
    except Exception as e:
        logger.error(f"Error processing PDF file {key}: {str(e)}")
        raise

def process_docx_as_text(bucket, key, config):
    """Fallback: Extract text from DOCX using ZIP structure"""
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        
        # Download document
        document_content = download_document(bucket, key)
        
        # Extract text from document.xml
        text_content = []
        
        with zipfile.ZipFile(BytesIO(document_content)) as docx_zip:
            # Get main document content
            with docx_zip.open('word/document.xml') as xml_file:
                tree = ET.parse(xml_file)
                root = tree.getroot()
                
                # Extract all text elements
                for elem in root.iter():
                    if elem.tag.endswith('}t'):  # Text elements
                        if elem.text:
                            text_content.append(elem.text)
        
        # Join all text with spaces
        full_text = ' '.join(text_content)
        
        # Apply redaction rules
        processed_text, redacted = apply_redaction_rules(full_text, config)
        
        # Save as text file (change extension to .txt)
        text_key = key.rsplit('.', 1)[0] + '.txt'
        
        # Upload processed document as text
        upload_processed_document(text_key, processed_text.encode('utf-8'), 
                                {'redacted': str(redacted), 'converted_from': 'docx', 'fallback_method': 'true'}, config)
        
        logger.info(f"Processed DOCX as text using fallback method: {key} -> {text_key}")
        return True
        
    except Exception as e:
        logger.error(f"Error in DOCX text extraction fallback for {key}: {str(e)}")
        raise

def process_docx_file(bucket, key, config):
    """Process DOCX files - convert to text and redact"""
    # Always use text extraction approach for consistency
    return process_docx_as_text(bucket, key, config)

def process_xlsx_file(bucket, key, config):
    """Process XLSX files - extract text and output as redacted text file"""
    if not load_workbook:
        raise ImportError("openpyxl library not available")
    
    try:
        # Download document
        document_content = download_document(bucket, key)
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx') as temp_file:
            temp_file.write(document_content)
            temp_file.flush()
            
            # Open workbook
            workbook = load_workbook(temp_file.name)
            
            # Extract all text content
            all_text = []
            
            # Process all worksheets
            for sheet in workbook.worksheets:
                sheet_text = [f"Sheet: {sheet.title}"]
                for row in sheet.iter_rows():
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            row_values.append(str(cell.value))
                    if row_values:
                        sheet_text.append('\t'.join(row_values))
                if len(sheet_text) > 1:  # Has content beyond title
                    all_text.append('\n'.join(sheet_text))
            
            # Combine all text
            full_text = '\n\n'.join(all_text)
            
            # Apply redaction rules
            processed_text, redacted = apply_redaction_rules(full_text, config)
            
            # Save as text file (change extension to .txt)
            text_key = key.rsplit('.', 1)[0] + '.txt'
            
            # Upload processed document as text
            upload_processed_document(text_key, processed_text.encode('utf-8'), 
                                    {'redacted': str(redacted), 'converted_from': 'xlsx'}, config)
            
            return True
            
    except Exception as e:
        logger.error(f"Error processing XLSX file {key}: {str(e)}")
        raise

def quarantine_document(bucket, key, reason):
    """Move document to quarantine bucket with retry logic"""
    copy_source = {'Bucket': bucket, 'Key': key}
    quarantine_key = f"quarantine/{key}"
    
    def _quarantine():
        s3.copy_object(
            CopySource=copy_source,
            Bucket=QUARANTINE_BUCKET,
            Key=quarantine_key,
            ServerSideEncryption='AES256',
            Metadata={
                'quarantine-reason': reason[:255],  # S3 metadata value limit
                'original-bucket': bucket,
                'original-key': key
            }
        )
    
    try:
        exponential_backoff_retry(_quarantine)
        logger.info(f"Document quarantined: s3://{QUARANTINE_BUCKET}/{quarantine_key}, reason: {reason}")
    except Exception as e:
        logger.error(f"Error quarantining document after {MAX_RETRIES} attempts: {str(e)}")
        raise